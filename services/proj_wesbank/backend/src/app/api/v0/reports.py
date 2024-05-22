import http
from io import BytesIO
import json

import requests
from datetime import datetime


from loguru import logger
import pandas as pd
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
import numpy as np
import traceback
from .helpers import  set_column_widths, write_filter_parameters,genrate_repair_count_df
from psycopg2 import sql
from .helpers import (
  
    last_twelve_months
)
from .db_config import exc_qrs_get_dfs_raw,  return_connection
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from .form_class import FormValues
from .auth import validate_token

reports_router = APIRouter()


# Gets the last log-in counts and times for wesbank auth0 tenant
@reports_router.post('/get_last_logged_users_logs', description='loggedin users report')
def get_last_loggin_report():
    conn = http.client.HTTPSConnection("fleet-analytics.eu.auth0.com")

    payload = "{\"client_id\":\"tccZOEAioj9FEWlbtzCBUAUUCdPtmEth\",\"client_secret\":\"wQJ3KxrUYpAkEBsevgk2qnuHpCRcl9olvXF95sI6l51tm7uIKrwQWN0PF0BWbfax\",\"audience\":\"https://fleet-analytics.eu.auth0.com/api/v2/\",\"grant_type\":\"client_credentials\"}"

    headers = { 'content-type': "application/json" }

    conn.request("POST", "/oauth/token", payload, headers)

    res = conn.getresponse()
    data = res.read().decode("utf-8")

    print(data)
    # quit()
    # Set your Auth0 domain and API token here
    AUTH0_DOMAIN = 'fleet-analytics.eu.auth0.com'
    API_URL = f'https://{AUTH0_DOMAIN}/api/v2/users'

    API_TOKEN = json.loads(data)['access_token']
    print(API_TOKEN)
    headers = {
        'Authorization': f'Bearer {API_TOKEN}',
        'Content-Type': 'application/json'
    }

    # Removed the query parameter that limited results to users who have logged in
    params = {
        'search_engine': 'v3',
        'fields': 'user_id,email,last_login,logins_count',
        'include_fields': 'true',
        'per_page': 100,  # Max number of users per page. Adjust as needed.
        'page': 0  # Start with the first page
    }

    all_users = []

    while True:
        response = requests.get(API_URL, headers=headers, params=params)
        if response.status_code == 200:
            users = response.json()
            if not users:
                break  # Exit the loop if no more users are returned
            for user in users:
                if user.get("last_login"):
                    last_login_datetime = datetime.strptime(user.get("last_login"), '%Y-%m-%dT%H:%M:%S.%fZ')
                    formatted_last_login = last_login_datetime.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    formatted_last_login = 'Never logged in'
                all_users.append({
        
                    'Email': user.get("email"),
                    'Last Login': formatted_last_login,
                    'Login Count': user.get("logins_count")
                })
            params['page'] += 1  # Move to the next page
        else:
            print('Failed to fetch users:', response.text)
            break

        # Convert the list of dictionaries to a DataFrame
        df = pd.DataFrame(all_users)
        template_path = "/app/app/reports_templates/users_report.xlsx"
        wb = load_workbook(template_path)
        buffer = BytesIO()
        dataframe_sheets = [
            {"sheet": "Users Report", "df": df, "start_row": 7},
          
        ]
                # write dfs to their sheets
        for sheet in dataframe_sheets:
            if sheet["sheet"] not in wb.sheetnames:
                wb.create_sheet(title=sheet["sheet"])
            ws = wb[sheet["sheet"]]
            start_row = sheet["start_row"]  #
            start_column = 1 

            # Convert DataFrame to rows
            rows = dataframe_to_rows(sheet["df"], index=False, header=True)

            # Write each row to the worksheet starting from the specific row and column
            for r_idx, row in enumerate(rows, start=start_row):
                for c_idx, value in enumerate(row, start=start_column):
                    ws.cell(row=r_idx, column=c_idx, value=value)       

        set_column_widths(wb)
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        print('xxx', df)
   

        # except Exception as error:
        #     logger.error(f"Error occurred: {error}")
        #     raise HTTPException(status_code=500, detail=f"Error Details: {traceback.format_exc()}")
        return StreamingResponse(
            BytesIO(buffer.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=data.csv"},)
# auth0 user logins report call



####################################  DRIVING EVENTS REPORT ####################################
## filter by vehicleregs


@reports_router.post(
    "/get_driving_events_report",
    description="Driving Events report",
    
)
def get_driving_events_report(formValues: dict, user: dict = Depends(validate_token)):
    try:
        form = FormValues(formValues)      
        driving_events_query = sql.SQL(
            """COPY(
            select division, vehiclereg,
            event_description,             
            event_start_date || ' ' || event_start_time as event_date,
            f_start_street || ' ' || f_start_suburb || ' ' || f_start_region as event_region,
                 fleet_no,veh_type_map,
            veh_model_map, Julian_month, asset_name,
            start_lat, start_lon,
            event_key 
            from fleet.driving_events
            where vehiclereg = any({registrations}) and  division = any({division}) and   (julian_month BETWEEN {julian_from} AND {julian_to} )
             ) TO STDOUT WITH CSV HEADER"""
        ).format(
             division=form.division,
            registrations=form.registrations,
            julian_from=form.julStartMonth,
            julian_to=form.julEndMonth,
          
        )

        driving_events_stats_query = sql.SQL(
            """COPY(
         select distinct vehiclereg,fleet_no,division, branch,veh_type_map, veh_make_map,
            sum ( case when event_description = 'Severe Impact' or event_description = 'Moderate Impact' or event_description = 'Impact' then 1 else 0 end) as impact_count,
            sum ( case when event_description = 'Acceleration' then 1 else 0 end) as acceleration_count,
            sum ( case when event_description = 'Braking' then 1 else 0 end) as braking_count, 
            sum ( case when event_description = 'Speeding' then 1 else 0 end) as overspeeding_count, 
            sum ( case when event_description = 'Cornering' then 1 else 0 end) as harsh_cornering_count

            from fleet.driving_events
            where vehiclereg = any({registrations}) and  (julian_month BETWEEN {julian_from} AND {julian_to} )
            group by vehiclereg,fleet_no,division,branch,veh_type_map,veh_make_map) TO STDOUT WITH CSV HEADER"""
        ).format(
            registrations=form.registrations,
            julian_from=form.julStartMonth,
            julian_to=form.julEndMonth,
        )

        results = exc_qrs_get_dfs_raw([driving_events_query, driving_events_stats_query])
        driving_events_query = results[0]
        driving_events_stats_query = results[1]      

        ###### Testing using the test_temp file
        template_path = "/app/app/reports_templates/driving_events_reports.xlsx"
        wb = load_workbook(template_path)

        buffer = BytesIO()
        dataframe_sheets = [
            {"sheet": "Driving Events", "df": driving_events_query, "start_row": 7},
            {"sheet": "Driving Events Counts", "df": driving_events_stats_query, "start_row": 7},
          
        ]
        # write dfs to their sheets
        for sheet in dataframe_sheets:
            if sheet["sheet"] not in wb.sheetnames:
                wb.create_sheet(title=sheet["sheet"])
            ws = wb[sheet["sheet"]]
            start_row = sheet["start_row"]  #
            start_column = 1 

            # Convert DataFrame to rows
            rows = dataframe_to_rows(sheet["df"], index=False, header=True)

            # Write each row to the worksheet starting from the specific row and column
            for r_idx, row in enumerate(rows, start=start_row):
                for c_idx, value in enumerate(row, start=start_column):
                    ws.cell(row=r_idx, column=c_idx, value=value)       

        set_column_widths(wb)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        # with pd.ExcelWriter(buffer) as writer:
        #     results.to_excel(writer, index=False)
    except Exception as error:
        logger.error(f"Error occurred: {error}")
        raise HTTPException(status_code=500, detail=f"Error Details: {traceback.format_exc()}")
    return StreamingResponse(
        BytesIO(buffer.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=data.csv"},)


####################################  COST PER SUPPLIER/COMPONENT REPORT ####################################
## one sheet, all transaction for a specific supplier/component.
## filter by vehicleregs, components and suppliers
@reports_router.post(
    "/get_per_supplier_or_component_report",
    description="Per supplier or component report",
    
)
def get_per_supplier_or_component_report(formValues: dict, user: dict = Depends(validate_token)):
    try:
        branches = formValues["branch"]
        providers = formValues["suppliers"]
        components = formValues["components"]
        registrations = list(map(lambda x: x["vehiclereg"], formValues["registrations"]))
        julian_month = formValues["julMonth"]
        from_month = formValues["julStartMonth"]
        to_month = formValues["julEndMonth"]
        miles_query = sql.SQL(
            """COPY(
            select * from fleet.maintenance where
            --branch = any({branches})
            --and
            serviceprovider = any({providers})
            and 
            mapping = any({components})
            and 
            vehiclereg = any({registrations})
            and 
            (julian_month between {from_month} and {to_month})
            order by transdate desc) TO STDOUT WITH CSV HEADER"""
        ).format(
            branches=sql.Literal(branches),
            providers=sql.Literal(providers),
            components=sql.Literal(components),
            registrations=sql.Literal(registrations),
            julian_month=sql.Literal(julian_month),
            from_month=sql.Literal(from_month),
            to_month=sql.Literal(to_month),
        )

        results_df = exc_qrs_get_dfs_raw([miles_query])[0]
        buffer = BytesIO()
        wb = Workbook()
        sheet = wb.active

        # Write header row
        header = [results_df.columns.values.tolist()]
        for row in header:
            sheet.append(row)
        # Write data
        for index, row in results_df.iterrows():
            sheet.append(row.tolist())
        set_column_widths(wb)

        wb.save(buffer)
        buffer.seek(0)
    except Exception as error:
        logger.error(f"Error occurred: {error}")
        raise HTTPException(status_code=500, detail=f"Error Details: {traceback.format_exc()}")
    return StreamingResponse(
        BytesIO(buffer.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=data.csv"},
    )


####################################  Detailed Usage Report ####################################
## Single sheet, dump of trip_data_daily
## filter by vehicleregs, julian_month
@reports_router.post(
    "/sho002_get_detailed_usage",
    description="Detailed Usage Report",
    
)
def sho002_get_detailed_usage(formValues: dict, user: dict = Depends(validate_token)):
    try:
        registrations = list(map(lambda x: x["vehiclereg"], formValues["registrations"]))
        julian_month = formValues["julMonth"]
        from_month = formValues["julStartMonth"]
        to_month = formValues["julEndMonth"]
        query = sql.SQL(
            """COPY(
                    SELECT
                        date,
                        fleet_no,
                        vehiclereg,
                        distance,
                        division,
                        branch,
                        veh_make_map as make,
                        veh_model_map as model,
                        driving_time_percentage,
                        standing_time_percentage,
                        idle_time_percentage,
                        total_driving_time,
                        total_standing_time,
                        total_duration,
                        total_idle_time,
                        parking_time,
                        total_engine_hours,
                        average_speed,
                        max_speed,
                        number_of_trips
                    FROM
                        fleet.trip_data_daily
                    WHERE 
                        (julian_month between {from_month} and {to_month})
                        AND vehiclereg = ANY({registrations})
                    ORDER BY date ASC
        )TO STDOUT WITH CSV HEADER
    """
        ).format(
            julian_month=sql.Literal(julian_month),
            registrations=sql.Literal(registrations),
            from_month=sql.Literal(from_month),
            to_month=sql.Literal(to_month),
        )
        # print(query)
        response = exc_qrs_get_dfs_raw([query])[0]
        buffer = BytesIO()
        wb = Workbook()
        sheet = wb.active

        # Write header row
        header = [response.columns.values.tolist()]
        for row in header:
            sheet.append(row)
        # Write data
        for index, row in response.iterrows():
            sheet.append(row.tolist())
        set_column_widths(wb)

        wb.save(buffer)
        buffer.seek(0)
    except Exception as error:
        logger.error(f"Error occurred: {error}")
        raise HTTPException(status_code=500, detail=f"Error Details: {traceback.format_exc()}")
    return StreamingResponse(
        BytesIO(buffer.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=data.csv"},
    )


####################################  COST PER ODO BAND Report ####################################
## Single sheet, shows vehciles' spend per component per odo band
## filter by vehicleregs, components
@reports_router.post(
    "/odo_band_per_component_report",
    description="Detailed Usage Report",
    
)
def odo_band_per_component_report(formValues: dict, user: dict = Depends(validate_token)):
    try:
        registrations = list(map(lambda x: x["vehiclereg"], formValues["registrations"]))
        components = formValues["components"]
        query = sql.SQL(
            """COPY(
                SELECT 
        vehiclereg,
        fleet_no,
        veh_type_map,
        mapping,
        SUM(CASE WHEN work_order_distance::float BETWEEN 0 AND 100000 THEN 1 ELSE 0 END) AS "0-100000km_count",
        SUM(CASE WHEN work_order_distance::float BETWEEN 0 AND 100000 THEN amount ELSE 0 END) AS "0-100000km_total",
        
        SUM(CASE WHEN work_order_distance::float > 100000 AND work_order_distance::float <= 200000 THEN 1 ELSE 0 END) AS "100001-200000km_count",
        SUM(CASE WHEN work_order_distance::float > 100000 AND work_order_distance::float <= 200000 THEN amount ELSE 0 END) AS "100001-200000km_total",

        SUM(CASE WHEN work_order_distance::float > 200000 AND work_order_distance::float <= 300000 THEN 1 ELSE 0 END) AS "200001-300000km_count",
        SUM(CASE WHEN work_order_distance::float > 200000 AND work_order_distance::float <= 300000 THEN amount ELSE 0 END) AS "200001-300000km_total",

        SUM(CASE WHEN work_order_distance::float > 300000 AND work_order_distance::float <= 400000 THEN 1 ELSE 0 END) AS "300001-400000km_count",
        SUM(CASE WHEN work_order_distance::float > 300000 AND work_order_distance::float <= 400000 THEN amount ELSE 0 END) AS "300001-400000km_total",

        SUM(CASE WHEN work_order_distance::float > 400000 AND work_order_distance::float <= 500000 THEN 1 ELSE 0 END) AS "400001-500000km_count",
        SUM(CASE WHEN work_order_distance::float > 400000 AND work_order_distance::float <= 500000 THEN amount ELSE 0 END) AS "400001-500000km_total",

        SUM(CASE WHEN work_order_distance::float > 500000 AND work_order_distance::float <= 600000 THEN 1 ELSE 0 END) AS "500001-600000km_count",
        SUM(CASE WHEN work_order_distance::float > 500000 AND work_order_distance::float <= 600000 THEN amount ELSE 0 END) AS "500001-600000km_total",

        SUM(CASE WHEN work_order_distance::float > 600000 AND work_order_distance::float <= 700000 THEN 1 ELSE 0 END) AS "600001-700000km_count",
        SUM(CASE WHEN work_order_distance::float > 600000 AND work_order_distance::float <= 700000 THEN amount ELSE 0 END) AS "600001-700000km_total",

        SUM(CASE WHEN work_order_distance::float > 700000 AND work_order_distance::float <= 800000 THEN 1 ELSE 0 END) AS "700001-800000km_count",
        SUM(CASE WHEN work_order_distance::float > 700000 AND work_order_distance::float <= 800000 THEN amount ELSE 0 END) AS "700001-800000km_total",

        SUM(CASE WHEN work_order_distance::float > 800000 AND work_order_distance::float <= 900000 THEN 1 ELSE 0 END) AS "800001-900000km_count",
        SUM(CASE WHEN work_order_distance::float > 800000 AND work_order_distance::float <= 900000 THEN amount ELSE 0 END) AS "800001-900000km_total",

        SUM(CASE WHEN work_order_distance::float > 900000 AND work_order_distance::float <= 1000000 THEN 1 ELSE 0 END) AS "900001-1000000km_count",
        SUM(CASE WHEN work_order_distance::float > 900000 AND work_order_distance::float <= 1000000 THEN amount ELSE 0 END) AS "900001-1000000km_total",
        
        SUM(CASE WHEN work_order_distance::float > 900000 AND work_order_distance::float <= 1000000 THEN 1 ELSE 0 END) AS "1000001-1100000km_count",
        SUM(CASE WHEN work_order_distance::float > 900000 AND work_order_distance::float <= 1000000 THEN amount ELSE 0 END) AS "1000001-1100000km_total",
            
        SUM(CASE WHEN work_order_distance::float > 900000 AND work_order_distance::float <= 1000000 THEN 1 ELSE 0 END) AS "1100001-1200000km_count",
        SUM(CASE WHEN work_order_distance::float > 900000 AND work_order_distance::float <= 1000000 THEN amount ELSE 0 END) AS "1100001-1200000km_total"
        FROM 
            fleet.maintenance
        WHERE vehiclereg = ANY({registrations})
        and mapping = any({components})
        GROUP BY 
        vehiclereg, veh_type_map, fleet_no, mapping

        order by vehiclereg

        )TO STDOUT WITH CSV HEADER
    """
        ).format(
            registrations=sql.Literal(registrations), components=sql.Literal(components)
        )
        # print(query)
        response = exc_qrs_get_dfs_raw([query])[0]
        buffer = BytesIO()
        wb = Workbook()
        sheet = wb.active

        # Write header row
        header = [response.columns.values.tolist()]
        for row in header:
            sheet.append(row)
        # Write data
        for index, row in response.iterrows():
            sheet.append(row.tolist())
        # add filters sheet
        ws = wb.create_sheet("Filters")
        write_filter_parameters(wb, "Filters", 1, formValues)
        set_column_widths(wb)

        wb.save(buffer)
        buffer.seek(0)
    except Exception as error:
        logger.error(f"Error occurred: {error}")
        raise HTTPException(status_code=500, detail=f"Error Details: {traceback.format_exc()}")
    return StreamingResponse(
        BytesIO(buffer.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=data.csv"},
    )


####################################  MONTHLY ORDERS REPORT ####################################
## two sheets, one contains all miles info, the other all the order info.
## filter by vehicleregs, components and suppliers

## the orders query currently only filters on branches as there are still many historical orders that does not have correct mapping or are for vehicles that we do not have on the fleetlist


@reports_router.post(
    "/get_monthly_orders_report",
    description="Monthly orders report",
    
)
def get_monthly_orders_report(formValues: dict, user: dict = Depends(validate_token)):
    try:

        branches = formValues["branch"]
        providers = formValues["suppliers"]
        components = formValues["components"]
        registrations = list(map(lambda x: x["vehiclereg"], formValues["registrations"]))
        julian_month = formValues["julMonth"]
        from_month = formValues["julStartMonth"]
        to_month = formValues["julEndMonth"]
        miles_query = sql.SQL(
            """COPY(
            select * from fleet.maintenance where
            branch = any({branches})
            and
            serviceprovider = any({providers})
            and 
            mapping = any({components})
            and 
            vehiclereg = any({registrations})
            and 
            (julian_month between {from_month} and {to_month})
            order by transdate desc) TO STDOUT WITH CSV HEADER"""
        ).format(
            branches=sql.Literal(branches),
            providers=sql.Literal(providers),
            components=sql.Literal(components),
            registrations=sql.Literal(registrations),
            julian_month=sql.Literal(julian_month),
            from_month=sql.Literal(from_month),
            to_month=sql.Literal(to_month),
        )

        orders_query = sql.SQL(
            """COPY(
            select * from fleet.orders where
            branch = any({branches})
            --and
            --serviceprovider = any({providers})
            --and 
            --mapping = any({components})
            --and 
            --vehiclereg = any({registrations})
            and 
            (julian_month between {from_month} and {to_month})
            order by date desc) TO STDOUT WITH CSV HEADER"""
        ).format(
            branches=sql.Literal(branches),
            providers=sql.Literal(providers),
            components=sql.Literal(components),
            registrations=sql.Literal(registrations),
            julian_month=sql.Literal(julian_month),
            from_month=sql.Literal(from_month),
            to_month=sql.Literal(to_month),
        )

        # get the monthly kms per asset
        usage_query = sql.SQL(
            """COPY(
            select * from fleet.trip_data_monthly where
            vehiclereg = any({registrations})
            and
            (julian_month between {from_month} and {to_month})

            ) TO STDOUT WITH CSV HEADER"""
        ).format(
            registrations=sql.Literal(registrations),
            julian_month=sql.Literal(julian_month),
            from_month=sql.Literal(from_month),
            to_month=sql.Literal(to_month),
        )

        results = exc_qrs_get_dfs_raw([miles_query, orders_query, usage_query])
        miles_df = results[0]
        orders_df = results[1]
        usage_df = results[2]
        invoice_totals_per_supplier = (
            miles_df.groupby("serviceprovider")["amount"]
            .agg(sum_amount="sum", count_amount="count")  # Directly set column names in agg
            .reset_index()
            .sort_values(by="sum_amount", ascending=False)  # Use the new column name
            .reset_index(drop=True)
        )
        invoice_totals_per_map = (
            miles_df.groupby("mapping")["amount"]
            .agg(sum_amount="sum", count_amount="count")  # Directly set column names in agg
            .reset_index()
            .sort_values(by="sum_amount", ascending=False)  # Use the new column name
            .reset_index(drop=True)
        )
        # print(usage_df)
        # group miles df and merge usage into it
        cost_and_usage_totals_df = (
            pd.merge(
                (
                    miles_df.groupby(["vehiclereg", "fleet_no"])["amount"]
                    .agg(
                        sum_amount="sum", count_amount="count"
                    )  # Directly set column names in agg
                    .reset_index()
                    .sort_values(
                        by="sum_amount", ascending=False
                    )  # Use the new column name
                    .reset_index(drop=True)
                ),
                usage_df[["vehiclereg", "distance"]],
                on="vehiclereg",
                how="left",
            )
            .fillna(0)
            .replace([np.inf, -np.inf], 0)
        )
        # determine overall cpk for the month
        cost_and_usage_totals_df["cpk"] = (
            (
                cost_and_usage_totals_df["sum_amount"]
                / cost_and_usage_totals_df["distance"]
                * 100
            )
            .fillna(0)
            .replace([np.inf, -np.inf], 0)
        )
        # print(cost_and_usage_totals_df)

        template_path = "/app/app/reports_templates/monthly_orders_report.xlsx"
        wb = load_workbook(template_path)
        ws = wb.active  # Select the default worksheet
        buffer = BytesIO()
        dataframe_sheets = [
            {"sheet": "Miles", "df": miles_df, "start_row": 7},
            {"sheet": "Orders", "df": orders_df, "start_row": 7},
            {
                "sheet": "Repair Counts",
                "df": genrate_repair_count_df(miles_df),
                "start_row": 7,
            },
        ]
        # write dfs to their sheets
        for sheet in dataframe_sheets:
            ws = wb[sheet["sheet"]]
            start_row = sheet["start_row"]  # Define the start row for each sheet
            start_column = 1  # Define the start column, 1 corresponds to 'A'

            # Convert DataFrame to rows
            rows = dataframe_to_rows(sheet["df"], index=False, header=True)

            # Write each row to the worksheet starting from the specific row and column
            for r_idx, row in enumerate(rows, start=start_row):
                for c_idx, value in enumerate(row, start=start_column):
                    ws.cell(row=r_idx, column=c_idx, value=value)
        # write summary sheet
        ws = wb["Summary"]
        startRow = 7
        for index, r in invoice_totals_per_supplier.iterrows():
            ws.cell(row=startRow, column=2).value = r.serviceprovider
            ws.cell(row=startRow, column=3).value = r.sum_amount
            ws.cell(row=startRow, column=4).value = r.count_amount
            startRow += 1

        startRow = 7
        for index, r in invoice_totals_per_map.iterrows():
            ws.cell(row=startRow, column=6).value = r.mapping
            ws.cell(row=startRow, column=7).value = r.sum_amount
            ws.cell(row=startRow, column=8).value = r.count_amount
            startRow += 1

        startRow = 7
        for index, r in cost_and_usage_totals_df.iterrows():
            ws.cell(row=startRow, column=10).value = r.vehiclereg
            ws.cell(row=startRow, column=11).value = r.fleet_no
            ws.cell(row=startRow, column=12).value = r.sum_amount
            ws.cell(row=startRow, column=13).value = r.count_amount
            ws.cell(row=startRow, column=14).value = round(r.cpk)
            ws.cell(row=startRow, column=15).value = round(r.distance)
            startRow += 1

        ##loop through divisions and write each division in the next row starting at filterStartRow
        write_filter_parameters(wb, "Filters", 5, formValues)
        set_column_widths(wb)

        # Save your workbook to a buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        # with pd.ExcelWriter(buffer) as writer:
        #     results.to_excel(writer, index=False)
    except Exception as error:
        logger.error(f"Error occurred: {error}")
        raise HTTPException(status_code=500, detail=f"Error Details: {traceback.format_exc()}")
    return StreamingResponse(
        BytesIO(buffer.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=data.csv"},
    )


####################################  FLEETLIST ####################################
## One sheet, contains fleet.fleetlist, without fields used for backend
## filter by vehicleregs


@reports_router.post(
    "/sho002_get_fleetlist_report",
    description="Fleetlist",
    
)
def sho002_get_fleetlist_report(formValues: dict, user: dict = Depends(validate_token)):
    try:
        form = FormValues(formValues)
        query = sql.SQL(
            """COPY(
        SELECT
            vehiclereg,
            fleet_no,
            deal_number,
            contract_type,
            division,
            branch,
            chassis_no,
            mm_code,
            vehicle_cat,
            description,
            new_used,
            maint_plan_cost,
            contract_mileage,
            make,
            veh_model_map,
            veh_type_map,
            date_of_first_reg,
            months_remaining,
            veh_lic_exp,
            contract_start,
            contract_end,
            last_odo,
            last_odo_date
        FROM
            fleet.fleetlist
        WHERE
            vehiclereg = ANY({vehiclereg})
        )TO STDOUT WITH CSV HEADER
                        """
        ).format(vehiclereg=form.registrations)

        response = exc_qrs_get_dfs_raw([query])[0]
        buffer = BytesIO()
        wb = Workbook()
        sheet = wb.active

        # Write header row
        header = [response.columns.values.tolist()]
        for row in header:
            sheet.append(row)
        # Write data
        for index, row in response.iterrows():
            sheet.append(row.tolist())
        set_column_widths(wb)

        wb.save(buffer)
        buffer.seek(0)
    except Exception as error:
        logger.error(f"Error occurred: {error}")
        raise HTTPException(status_code=500, detail=f"Error Details: {traceback.format_exc()}")
    return StreamingResponse(
        BytesIO(buffer.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=data.csv"},
    )


@reports_router.post(
    "/sho002_get_monthly_vehicles_report",
    description="monthly vehicle file",
  
    
)
def monthly_vehicles_report(formValues: dict, user: dict = Depends(validate_token)):
    try:
        form = FormValues(formValues)
        query = sql.SQL(
            """ COPY(
                     select * from fleet.fleetlist 
                     where date_of_first_reg between  {julian_from} AND {julian_to}           
        )TO STDOUT WITH CSV HEADER """
        ).format(
                julian_from=form.julStartMonth,
            julian_to=form.julEndMonth,)
        response = exc_qrs_get_dfs_raw([query])
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error Details: {traceback.format_exc()}")
    return response[0].to_dict("records")

@reports_router.post("/external_fleetlist_report")
def post_write_to_spreadsheet(formValues: dict, user: dict = Depends(validate_token)):
    try:
        form = FormValues(formValues)
        write_to_file_external_fleetlist(form)

        with open("/app/app/reports_templates/external_fletlist_report.xlsx", "rb") as file:
            contents = file.read()
    except Exception as e:
        print(f"An error occurred: {e}")
        raise HTTPException(status_code=500, detail=f"Error Details: {traceback.format_exc()}")
    return StreamingResponse(
        BytesIO(contents),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment;filename=external_fletlist_report.xlsx"
        },
    )


@reports_router.post("/external_monthly_vehicles_report")
def monthly_vehicles_report(from_date: str, to_date: str, user: dict = Depends(validate_token)):
    try:
        with open("/app/app/reports_templates/monthly_vehicles_report.xlsx", "rb") as file:
            contents = file.read()
    except Exception as e:
        print(f"An error occurred: {e}")
        raise HTTPException(status_code=500, detail=f"Error Details: {traceback.format_exc()}")
    return StreamingResponse(
        BytesIO(contents),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment;filename=monthly_vehicles_report.xlsx"
        },
    )


@reports_router.post(
    "/get_invoices",
    description="Get invoices per model group for a component map",
 
)
def get_component_map_invoices(
    formValues: dict, user: dict = Depends(validate_token)
):
    try:
        form = FormValues(formValues)
        query_list = []
        invoice_query = sql.SQL(
            """COPY(with vehicleregs as (select vehiclereg from fleet.fleetlist where veh_model_map = any({models}) )
    select * from fleet.maintenance 
    where vehiclereg in (select vehiclereg from vehicleregs)
    and mapping = any({component_map})
    order by transdate desc
    )  TO STDOUT WITH CSV HEADER"""
        ).format(
           model=form.models,
         components=form.components,
        )
        query_list.append(invoice_query)
        # try:
        response_list = exc_qrs_get_dfs_raw(query_list)

        df = response_list[0]
        vehicle_df = df[df.vehiclereg == vehicle_reg]
    except Exception as e:
        print(f"An error occurred: {e}")
        raise HTTPException(status_code=500, detail=f"Error Details: {traceback.format_exc()}")
    return vehicle_df.to_dict("records")

@reports_router.post(
    "/sho002_get_fleetlist_raw_data",
    description="Get FleetList Raw Data",
)
def sho002_get_fleetlist_raw_data(formValues: dict, user: dict = Depends(validate_token)):
    try:
        form  = FormValues(formValues)        
        fleetlist_raw_data = sql.SQL(
            """COPY(
                                    select * from fleet.fleetlist_raw
                                    where vehiclereg = any({registrations}) 
                                
        )TO STDOUT WITH CSV HEADER """
        ).format(
            registrations=form.registrations,
            )

        response = exc_qrs_get_dfs_raw([fleetlist_raw_data])[0].to_dict("records")
    except Exception as e:
        print(f"An error occurred: {e}")
        raise HTTPException(status_code=500, detail=f"Error Details: {traceback.format_exc()}")
    return response


@reports_router.post(
    "/sho002_get_maint_raw_data",
    description="Get maintenance Raw Data",
)
def sho002_get_maint_raw_data(formValues: dict, user: dict = Depends(validate_token)):

    try:
        registrations = list(
            map(lambda x: x["vehiclereg"], formValues["registrations"])
        )
        components = formValues["components"]
        jul_month = formValues["julMonth"]
        maint_raw_data = sql.SQL(
            """COPY(
                    SELECT * FROM fleet.maintenance where
                    julian_month = {jul_month} 
                    and vehiclereg = any({registrations}) 
                    and mapping = any({components}) 
                    ORDER BY vehiclereg ASC, maintdescription ASC, transdate ASC, serviceprovider ASC, amount ASC                             
        )TO STDOUT WITH CSV HEADER """
        ).format(
            jul_month=sql.Literal(jul_month),
            registrations=sql.Literal(registrations),
            components=sql.Literal(components),
        )

        response = exc_qrs_get_dfs_raw([maint_raw_data])[0].to_dict("records")
    except Exception as e:
        print(f"An error occurred: {e}")
        raise HTTPException(status_code=500, detail=f"Error Details: {traceback.format_exc()}")
    return response

def write_to_file_accrual_report(
    from_date: str, to_date: str, division: str, branch: str, veh_type: str, user: dict = Depends(validate_token)
):
    try:
        con = return_connection()

        pivot_service_provider = f"""SELECT            
        serviceprovider,
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 1 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Jan 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 2 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Feb 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 3 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Mar 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 4 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Apr 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 5 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "May 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 6 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Jun 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 7 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Jul 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 8 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Aug 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 9 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Sep 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 10 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Oct 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 11 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Nov 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 12 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Dec 2023",
        SUM(amount) AS "Total"
        FROM fleet.maintenance
            GROUP BY 
                ROLLUP (serviceprovider)
            ORDER BY 
        serviceprovider"""

        shp_query = f"""SELECT * FROM fleet.maintenance where 
        lower(division) = '{division}' and lower(branch) = '{branch}' and 
        lower(veh_type_map) = '{veh_type} 'and transdate between '{from_date}' and '{to_date}'
        """
        pivot_branch = f"""SELECT            
        branch,
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 1 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Jan 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 2 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Feb 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 3 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Mar 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 4 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Apr 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 5 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "May 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 6 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Jun 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 7 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Jul 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 8 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Aug 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 9 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Sep 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 10 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Oct 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 11 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Nov 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 12 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Dec 2023",
        SUM(amount) AS "Total"
            FROM fleet.maintenance
            GROUP BY 
                ROLLUP (branch)
            ORDER BY 
        branch"""

        comparison_query = f"""	
        SELECT            
            veh_type_map,
            SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 1 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Jan 2023",
            SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 2 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Feb 2023",
            SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 3 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Mar 2023",
            SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 4 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Apr 2023",
            SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 5 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "May 2023",
            SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 6 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Jun 2023",
            SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 7 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Jul 2023",
            SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 8 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Aug 2023",
            SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 9 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Sep 2023",
            SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 10 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Oct 2023",
            SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 11 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Nov 2023",
            SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 12 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Dec 2023",
            SUM(amount) AS "Total"
        FROM fleet.maintenance
        WHERE
            veh_type_map in ('Passenger Car', 'Medium Commercial Vehicle','Light Commercial Vehicle', 'Heavy Commercial Vehicle' )

        GROUP BY 
            ROLLUP (veh_type_map) 
        ORDER BY 
            veh_type_map"""

        comparison_query_diff = f"""select	
            veh_type_map,
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 1 AND EXTRACT(YEAR FROM transdate) = 2022 THEN amount ELSE 0 END) AS "Jan 2022",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 2 AND EXTRACT(YEAR FROM transdate) = 2022 THEN amount ELSE 0 END) AS "Feb 2022",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 3 AND EXTRACT(YEAR FROM transdate) = 2022 THEN amount ELSE 0 END) AS "Mar 2022",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 4 AND EXTRACT(YEAR FROM transdate) = 2022 THEN amount ELSE 0 END) AS "Apr 2022",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 5 AND EXTRACT(YEAR FROM transdate) = 2022 THEN amount ELSE 0 END) AS "May 2022",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 6 AND EXTRACT(YEAR FROM transdate) = 2022 THEN amount ELSE 0 END) AS "Jun 2022",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 7 AND EXTRACT(YEAR FROM transdate) = 2022 THEN amount ELSE 0 END) AS "Jul 2022",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 8 AND EXTRACT(YEAR FROM transdate) = 2022 THEN amount ELSE 0 END) AS "Aug 2022",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 9 AND EXTRACT(YEAR FROM transdate) = 2022 THEN amount ELSE 0 END) AS "Sep 2022",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 10 AND EXTRACT(YEAR FROM transdate) = 2022 THEN amount ELSE 0 END) AS "Oct 2022",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 11 AND EXTRACT(YEAR FROM transdate) = 2022 THEN amount ELSE 0 END) AS "Nov 2022",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 12 AND EXTRACT(YEAR FROM transdate) = 2022 THEN amount ELSE 0 END) AS "Dec 2022",
        SUM(amount) AS "Total"
        FROM fleet.maintenance
        WHERE
            veh_type_map in ('Passenger Car', 'Medium Commercial Vehicle','Light Commercial Vehicle', 'Heavy Commercial Vehicle' )

        GROUP BY 
            ROLLUP (veh_type_map) 
        ORDER BY 
            veh_type_map"""

        comparison_query = pd.read_sql_query(
            comparison_query, con
        )  ## comparision query (passenger + commercial)
        comparison_query_diff = pd.read_sql_query(
            comparison_query_diff, con
        )  ## comparision query (passenger + commercial)

        pivot_sheet2 = pd.read_sql_query(pivot_branch, con)  ## branches
        pivot_sheet = pd.read_sql_query(pivot_service_provider, con)  ### servive providers
        ## shoprite transactions orders
        data_sheet4 = pd.read_sql_query(shp_query, con)
        ### accrual orders

    
        template_path = "/app/app/reports_templates/accrual_report_template.xlsx"
        workbook = openpyxl.load_workbook(template_path)

        sheet_name = "Data"
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")

        start_row = 2
        start_column = 1

        for col_num, header in enumerate(data_sheet4.columns, start_column):
            sheet.cell(row=2, column=col_num, value=header)

        for row_num, (_, row_data) in enumerate(data_sheet4.iterrows(), start_row):
            for col_num, header in enumerate(data_sheet4.columns, start_column):
                value = row_data[header]
                sheet.cell(row=row_num, column=col_num, value=value)

        workbook.save(template_path)
        # print("Data written to spreadsheet successfully!")

        template_path = "/app/app/reports_templates/accrual_report_template.xlsx"
        workbook = openpyxl.load_workbook(template_path)

        sheet_name = "Pivot"
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")

        start_row = 3
        start_column = 1

        for col_num, header in enumerate(pivot_sheet.columns, start_column):
            sheet.cell(row=2, column=col_num, value=header)

        for row_num, (_, row_data) in enumerate(pivot_sheet.iterrows(), start_row):
            for col_num, header in enumerate(pivot_sheet.columns, start_column):
                value = row_data[header]
                sheet.cell(row=row_num, column=col_num, value=value)

        workbook.save(template_path)
        # print("Data written to spreadsheet successfully!")

        template_path = "/app/app/reports_templates/accrual_report_template.xlsx"
        workbook = openpyxl.load_workbook(template_path)

        sheet_name = "Pivot"
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")

        start_row = 3
        start_column = 17

        start_column = 17

        for col_num, header in enumerate(pivot_sheet2.columns, start_column):
            sheet.cell(row=2, column=col_num, value=header)

        for row_num, (_, row_data) in enumerate(pivot_sheet2.iterrows(), start_row):
            for col_num, header in enumerate(pivot_sheet2.columns, start_column):
                value = row_data[header]
                sheet.cell(row=row_num, column=col_num, value=value)

        workbook.save(template_path)
        # print("Data written to spreadsheet successfully!")

        template_path = "/app/app/reports_templates/accrual_report_template.xlsx"
        workbook = openpyxl.load_workbook(template_path)

        sheet_name = "Comparison"
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")

        start_row = 2
        start_column = 1

        start_column = 1

        for col_num, header in enumerate(comparison_query.columns, start_column):
            sheet.cell(row=2, column=col_num, value=header)

        for row_num, (_, row_data) in enumerate(comparison_query.iterrows(), start_row):
            for col_num, header in enumerate(comparison_query.columns, start_column):
                value = row_data[header]
                sheet.cell(row=row_num, column=col_num, value=value)

        workbook.save(template_path)
        # print("Data written to spreadsheet successfully!")

        template_path = "/app/app/reports_templates/accrual_report_template.xlsx"
        workbook = openpyxl.load_workbook(template_path)

        sheet_name = "Comparison"
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")

        start_row = 14
        start_column = 1

        start_column = 1

        for col_num, header in enumerate(comparison_query_diff.columns, start_column):
            sheet.cell(row=2, column=col_num, value=header)

        for row_num, (_, row_data) in enumerate(
            comparison_query_diff.iterrows(), start_row
        ):
            for col_num, header in enumerate(
                comparison_query_diff.columns, start_column
            ):
                value = row_data[header]
                sheet.cell(row=row_num, column=col_num, value=value)

        workbook.save(template_path)
        print("Data written to spreadsheet successfully!")

    except Exception as e:
            print(f"An error occurred: {e}")
            raise HTTPException(status_code=500, detail=f"Error Details: {traceback.format_exc()}")

@reports_router.post("/accrual_report")
def generate_accrual_report(
    from_date: str, to_date: str, division: str, branch: str, veh_type: str, user: dict = Depends(validate_token)
):
    try:
        write_to_file_accrual_report(from_date, to_date, division, branch, veh_type)

        with open("/app/app/reports_templates/accrual_report_template.xlsx", "rb") as file:
            contents = file.read()
    except Exception as e:
        print(f"An error occurred: {e}")
        raise HTTPException(status_code=500, detail=f"Error Details: {traceback.format_exc()}")

    return StreamingResponse(
        BytesIO(contents),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=accrual_report.xlsx"},
    )


def write_to_file_external_fleetlist(division: str, branch: str, veh_type: str):
    try:

        con = return_connection()()
        sql_query = """SELECT * FROM fleet.fleetlist_external
        WHERE lower(division) = %(division)s AND lower(branch) = %(branch)s AND lower(veh_type_map) = %(veh_type)s"""

        df_sql = pd.read_sql_query(
            sql_query,
            con,
            params={"division": division, "branch": branch, "veh_type": veh_type},
        )

        # print(df_sql, "hhfdhfhfh")

        template_path = "/app/app/reports_templates/external_fletlist_report.xlsx"
        workbook = openpyxl.load_workbook(template_path)

        sheet_name = "external fleetlist"
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")

        start_row = 3
        start_column = 1

        for col_num, header in enumerate(df_sql.columns, start_column):
            sheet.cell(row=2, column=col_num, value=header)

        for row_num, (_, row_data) in enumerate(df_sql.iterrows(), start_row):
            for col_num, header in enumerate(df_sql.columns, start_column):
                value = row_data[header]
                sheet.cell(row=row_num, column=col_num, value=value)

        workbook.save(template_path)
            # print("Data written to spreadsheet successfully!")
    except Exception as e:
        print(f"An error occurred: {e}")
        raise HTTPException(status_code=500, detail=f"Error Details: {traceback.format_exc()}")


@reports_router.post(
    "/sho002_shoprite_accrual_report",
    description="get shoprite accrual report",
)
def sho002_shoprite_accrual_report(formValues: dict, user: dict = Depends(validate_token)):
    try:
        formValues = FormValues(formValues)
        report_table_query = sql.SQL(
            """COPY(
                        select 
                    *
                        from fleet.maintenance
                        where invoice_status = 'accrual' and  (julian_month BETWEEN {julian_from} AND {julian_to} )
                        and vehiclereg = any({registrations}) 
                            ) TO STDOUT WITH CSV HEADER"""
        ).format(
            julian_from=formValues.julStartMonth,
            julian_to=formValues.julEndMonth,
            registrations=formValues.registrations,
        )
        response = exc_qrs_get_dfs_raw([report_table_query])[0]
    except Exception as e:
        print(f"An error occurred: {e}")
        raise HTTPException(status_code=500, detail=f"Error Details: {traceback.format_exc()}")
    return response.to_dict("records")


@reports_router.post(
    "/sho002_get_report_fleetlist",
    description="get fleetlist for report download",
    
)
def sho002_get_report_fleetlist(
    formValues: dict, user: dict = Depends(validate_token)
):
    try:
        form = FormValues(formValues)
        type_query = sql.SQL(
            """COPY(
            select * from fleet.fleetlist   
            where  (julian_month BETWEEN {julian_from} AND {julian_to} )
                        and vehiclereg = any({registrations})                   
                            )TO STDOUT WITH CSV HEADER """
        ).format(
         julian_from=form.julStartMonth,
         julian_to=form.julEndMonth,
         registrations=form.registrations,
        )

        response = exc_qrs_get_dfs_raw([type_query])[0].to_dict("records")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error Details: {traceback.format_exc()}")
    return response


@reports_router.post(
    "/sho002_get_report_cost_per_supplier_per_month",
    description="get supplier cost per supplier per month",
    
)
def sho002_get_report_cost_per_supplier_per_month(
    formValues: dict, user: dict = Depends(validate_token)
):
    try:
        form = FormValues(formValues)
        type_query = sql.SQL(
            """COPY(
            select date_trunc('month', transdate) as month, serviceprovider,  sum(amount), division, branch, veh_type_map
            from fleet.maintenance 
            where (julian_month BETWEEN {julian_from} AND {julian_to} )
                        and vehiclereg = any({registrations})   
            group by month, serviceprovider, division, branch, veh_type_map
            order by month desc, serviceprovider                  
                            )TO STDOUT WITH CSV HEADER """
        ).format(
         julian_from=form.julStartMonth,
            julian_to=form.julEndMonth,
            registrations=form.registrations,
        )

        response = exc_qrs_get_dfs_raw([type_query])[0].to_dict("records")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error Details: {traceback.format_exc()}")
    return response


@reports_router.post(
    "/sho002_get_report_cost_per_component_per_month",
    description="get component cost per component per month",
    
)
def sho002_get_report_cost_per_component_per_month(
  formValues: dict, user: dict = Depends(validate_token)
):
    try:

        form = FormValues(formValues)
        type_query = sql.SQL(
            """COPY(
        select transdate, mapping, division, branch, veh_type_map from fleet.maintenance 
                            where 
                (julian_month BETWEEN {julian_from} AND {julian_to} )
                        and vehiclereg = any({registrations})          group by transdate, mapping, division, branch, veh_type_map 
        order by transdate, mapping asc           
                            )TO STDOUT WITH CSV HEADER """
        ).format(
            julian_from=form.julStartMonth,
            julian_to=form.julEndMonth,
            registrations=form.registrations,
        )

        response = exc_qrs_get_dfs_raw([type_query])[0].to_dict(
            "records"
        )  # DO NOT CHANGE THIS LINE
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error Details: {traceback.format_exc()}")
    return response


@reports_router.post(
    "/sho002_get_report_supplier_invoices_per_month",
    description="get supplier invoices per month",
    
)
def sho002_get_report_supplier_invoices_per_month(
formValues: dict, user: dict = Depends(validate_token)
):
    try:
        form = FormValues(formValues)        
        type_query = sql.SQL(
            """COPY(
        select *
        from fleet.maintenance 
        where 
        (julian_month BETWEEN {julian_from} AND {julian_to} )
                        and vehiclereg = any({registrations})          and serviceprovider = any({supplier})    
                            )TO STDOUT WITH CSV HEADER """
        ).format(
         julian_from=form.julStartMonth,
            julian_to=form.julEndMonth,
            registrations=form.registrations,
            supplier=form.supplier,
        )

        response = exc_qrs_get_dfs_raw([type_query])[0].to_dict(
        "records"
    )  # DO NOT CHANGE THIS LINE
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error Details: {traceback.format_exc()}")
    return response


@reports_router.post(
    "/sho002_get_report_component_invoices_per_month",
    description="get component invoices per month",
    
)
def sho002_get_report_component_invoices_per_month(
   formValues: dict, user: dict = Depends(validate_token)
):
    try:
        form = FormValues(formValues)
        type_query = sql.SQL(
            """COPY(
        select *
        from fleet.maintenance 
        where 
        (julian_month BETWEEN {julian_from} AND {julian_to} )
        and vehiclereg = any({registrations})            
        and mapping = any({component})         
                            )TO STDOUT WITH CSV HEADER """
        ).format(
            julian_from=form.julStartMonth,
            julian_to=form.julEndMonth,
            registrations=form.registrations,
            components=form.components,
        )

        response = exc_qrs_get_dfs_raw([type_query])[0]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error Details: {traceback.format_exc()}")
    return response.to_dict("records")  # DO NOT CHANGE THIS LINE


@reports_router.post(
    "/sho002_get_report_trip_data_per_month",
    description="get trip data per month",
    
)
def sho002_get_report_trip_data_per_month(formValues: dict, user: dict = Depends(validate_token)):
    try:

        form = FormValues(formValues)
        type_query = sql.SQL(
            """COPY(
                    select * from fleet.trip_data_cent
                    where   (julian_month BETWEEN {julian_from} AND {julian_to} )
                    order by date desc              
            )TO STDOUT WITH CSV HEADER """
        ).format(
           julian_from=form.julStartMonth,
            julian_to=form.julEndMonth,
        )

        response = exc_qrs_get_dfs_raw([type_query])[0]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error Details: {traceback.format_exc()}")
    return response.to_dict("records")  # DO NOT CHANGE THIS LINE


@reports_router.post(
    "/sho002_get_report_asset_invoices_per_month",
    description="get maintenance maps ",
    
)
def sho002_get_report_asset_invoices_per_month(
   formValues: dict, user: dict = Depends(validate_token)
):
    try:
        form = FormValues(formValues)
        type_query = sql.SQL(
            """COPY(
            select * from fleet.maintenance 
                where vehiclereg = any({registrations})  
                    and
                    (julian_month BETWEEN {julian_from} AND {julian_to} )
                order by transdate desc           
            )TO STDOUT WITH CSV HEADER """
        ).format(
          julian_from=form.julStartMonth,
            julian_to=form.julEndMonth,
            registrations=form.registrations,
        )

        response = (
            exc_qrs_get_dfs_raw([type_query])[0].fillna(0).to_dict("records")
        )  # DO NOT CHANGE THIS LINE
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error Details: {traceback.format_exc()}")
    return response


# shoprite checkers rebills report
@reports_router.post(
    "/shoprite_checkers_rebills_report",
    description="shoprite checkers rebills report",
    
)
def shoprite_checkers_rebills_report(user: dict = Depends(validate_token)):
    try:
        query = sql.SQL(
            """ COPY(
                     select * from fleet.accrual_orders
        )TO STDOUT WITH CSV HEADER """
        ).format()
        response = exc_qrs_get_dfs_raw([query])
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error Details: {traceback.format_exc()}")
    return response[0].to_dict("records")


@reports_router.post(
    "/sho002_get_usage_summary",
    description="Usage summary",
    
)
def sho002_get_usage_summary(formValues: dict, user: dict = Depends(validate_token)):
    try:
        branches = formValues["branch"]
        registrations = list(map(lambda x: x["vehiclereg"], formValues["registrations"]))
        julian_month = formValues["julMonth"]
        from_month = formValues["julStartMonth"]
        to_month = formValues["julEndMonth"]
        query = sql.SQL(
            """COPY(
        WITH branches AS(
        SELECT
            ftm.vehiclereg,
            ftm.fleet_no,
            ftm.veh_type_map,
            ftm.veh_model_map,
            ftm.distance,
            CASE
                WHEN AVG(ftm.distance) OVER (PARTITION BY ftm.veh_type_map) = 0 THEN 0
                ELSE ROUND((ftm.distance / AVG(ftm.distance) OVER (PARTITION BY ftm.veh_type_map)) * 100, 2)
            END AS pct_of_avg_type,
            CASE
                WHEN AVG(ftm.distance) OVER (PARTITION BY ftm.veh_model_map) = 0 THEN 0
                ELSE ROUND((ftm.distance / AVG(ftm.distance) OVER (PARTITION BY ftm.veh_model_map)) * 100, 2)
            END AS pct_of_avg_model
        FROM
            fleet.trip_data_monthly ftm
        WHERE
            (julian_month between {from_month} and {to_month})
            AND branch = ANY({branches}) 
        ),
        fullfleet AS(
            SELECT
                ftm.vehiclereg,
                CASE
                    WHEN AVG(ftm.distance) OVER (PARTITION BY ftm.veh_type_map) = 0 THEN 0
                    ELSE ROUND((ftm.distance / AVG(ftm.distance) OVER (PARTITION BY ftm.veh_type_map)) * 100, 2)
                END AS pct_of_avg_type,
                CASE
                    WHEN AVG(ftm.distance) OVER (PARTITION BY ftm.veh_model_map) = 0 THEN 0
                    ELSE ROUND((ftm.distance / AVG(ftm.distance) OVER (PARTITION BY ftm.veh_model_map)) * 100, 2)
                END AS pct_of_avg_model
        FROM
            fleet.trip_data_monthly ftm
        WHERE
            (julian_month between {from_month} and {to_month}))
        
    SELECT
        branches.vehiclereg,
        branches.fleet_no,
        branches.veh_type_map,
        branches.veh_model_map,
        branches.distance,
        branches.pct_of_avg_type AS pct_usage_vs_type_avg,
        branches.pct_of_avg_model AS pct_usage_vs_model_avg,
        fullfleet.pct_of_avg_type AS pct_usage_vs_fleet_type_avg,
        fullfleet.pct_of_avg_model AS pct_usage_vs_fleet_model_avg
    FROM branches
    JOIN fullfleet ON branches.vehiclereg = fullfleet.vehiclereg
    WHERE
        branches.vehiclereg = ANY({vehicle_list})
        )TO STDOUT WITH CSV HEADER
    """
        ).format(
            julian_month=sql.Literal(julian_month),
            vehicle_list=sql.Literal(registrations),
            branches=sql.Literal(branches),
            from_month=sql.Literal(from_month),
            to_month=sql.Literal(to_month),
        )
        response = exc_qrs_get_dfs_raw([query])[0]
        buffer = BytesIO()
        wb = Workbook()
        sheet = wb.active

        # Write header row
        header = [response.columns.values.tolist()]
        for row in header:
            sheet.append(row)
        # Write data
        for index, row in response.iterrows():
            sheet.append(row.tolist())
        set_column_widths(wb)

        wb.save(buffer)
        buffer.seek(0)
    except Exception as error:
        logger.error(f"Error occurred: {error}")
        raise HTTPException(status_code=500, detail=f"Error Details: {traceback.format_exc()}")
    return StreamingResponse(
        BytesIO(buffer.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=data.csv"},
    )

@reports_router.post("/contract_usage_twelve_month",
                     description = 'Gets 12 month contract usage report')
def get_twelve_month(formValues: dict, user: dict = Depends(validate_token)):

    form = FormValues(formValues)
    months = last_twelve_months()
    month_sum_cases = ",\n    ".join(
        [
            f"SUM(CASE WHEN month = '{month}' THEN distance ELSE 0 END) AS \"{month}\""
            for month in months
        ]
    )
    
    query = sql.SQL("""COPY(
    with monthly_totals as (
        SELECT
            vehiclereg, fleet_no, division, branch, veh_type_map, veh_make_map, veh_model_map,
            TO_CHAR(julian_month, 'YYYY-MM') AS month,
            ROUND(SUM(distance)) AS distance
        FROM
            fleet.trip_data_daily
        WHERE
            julian_month >= NOW() - INTERVAL '12 MONTHS'
            AND vehiclereg = ANY({reg})
        GROUP BY
            vehiclereg, fleet_no, division, branch, veh_type_map, veh_make_map, veh_model_map, month
    )
    SELECT
        monthly_totals.vehiclereg, monthly_totals.fleet_no, monthly_totals.division, monthly_totals.branch, monthly_totals.veh_type_map, monthly_totals.veh_make_map, monthly_totals.veh_model_map, fl.contract_mileage,
        {month_cases},
        SUM(distance) AS grand_total
    FROM
        monthly_totals
    LEFT JOIN fleet.fleetlist fl ON fl.vehiclereg = monthly_totals.vehiclereg
    GROUP BY
        monthly_totals.vehiclereg, monthly_totals.fleet_no, monthly_totals.division, monthly_totals.branch, monthly_totals.veh_type_map, monthly_totals.veh_make_map, monthly_totals.veh_model_map, fl.contract_mileage
    ORDER BY
        vehiclereg
) TO STDOUT WITH CSV HEADER
""").format(
    reg=form.registrations,
    month_cases=sql.SQL(month_sum_cases)
)

    result = exc_qrs_get_dfs_raw([query])[0]
    buffer = BytesIO()
    wb = Workbook()
    sheet = wb.active

        # Write header row
    header = [result.columns.values.tolist()]
    for row in header:
        sheet.append(row)
    # Write data
    for index, row in result.iterrows():
        sheet.append(row.tolist())
    set_column_widths(wb)

    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        BytesIO(buffer.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=data.csv"},
    )

@reports_router.post("/downtime_report",
                     description = 'Gets 12 month contract usage report')
def downtime_report(formValues: dict, user: dict = Depends(validate_token)):

    form = FormValues(formValues)
   
    query = sql.SQL("""COPY(
        SELECT 
            vehiclereg,
            fleet_no,
            branch,
            type,
            make,
            model,
            odo,
            supplier,
            reason,
            start_date,
            end_date,
            est_end_date
        FROM fleet.downtime
        WHERE vehiclereg = ANY({reg})
        AND start_date BETWEEN (SELECT jul_from_date FROM fleet.julian_cal WHERE selected_month = {from_date}) AND (SELECT jul_to_date FROM fleet.julian_cal WHERE selected_month = {to_date})
) TO STDOUT WITH CSV HEADER
""").format(
    reg=form.registrations,
    from_date=form.julStartMonth,
    to_date=form.julEndMonth
)

    result = exc_qrs_get_dfs_raw([query])[0]
    buffer = BytesIO()
    wb = Workbook()
    sheet = wb.active

        # Write header row
    header = [result.columns.values.tolist()]
    for row in header:
        sheet.append(row)
    # Write data
    for index, row in result.iterrows():
        sheet.append(row.tolist())
    set_column_widths(wb)

    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        BytesIO(buffer.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=data.csv"},
    )

@reports_router.post("/fleetcard_report",
                     description = 'Gets 12 month contract usage report')
def fleetcard_report(formValues: dict, user: dict = Depends(validate_token)):

    form = FormValues(formValues)
   
    query = sql.SQL("""COPY(
        SELECT 
            vehiclereg,
            fleet_no,
            vendor,
            area,
            transaction_date,
            transaction_cost,
            litres,
            transaction_type,
            transaction_number
        FROM fleet.fleet_card
        WHERE vehiclereg = ANY({reg})
        AND julian_month BETWEEN {from_date} AND {to_date}
) TO STDOUT WITH CSV HEADER
""").format(
    reg=form.registrations,
    from_date=form.julStartMonth,
    to_date=form.julEndMonth
)

    result = exc_qrs_get_dfs_raw([query])[0]
    buffer = BytesIO()
    wb = Workbook()
    sheet = wb.active

        # Write header row
    header = [result.columns.values.tolist()]
    for row in header:
        sheet.append(row)
    # Write data
    for index, row in result.iterrows():
        sheet.append(row.tolist())
    set_column_widths(wb)

    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        BytesIO(buffer.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=data.csv"},
    )