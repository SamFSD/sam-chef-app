import traceback
from psycopg2 import sql
from .db_config import exc_qrs_get_dfs_raw, return_connection
import pandas as pd
from typing import List
import openpyxl
from datetime import datetime
from fastapi import APIRouter,HTTPException
from openpyxl.utils import get_column_letter
from datetime import timedelta

##
#
# filter api arguments for fleet and branch selections
#
# if full fleet or all branches return uneditted sql snippets

helper_func = APIRouter()
def supplier_filter_check(supplier: str):
    supplier = supplier.lower().split(",")
    if "all_suppliers" in supplier:
        supplier_filter = sql.SQL("lower(serviceprovider)")
    else:
        supplier_filter = sql.SQL("any({supplier})").format(
            supplier=sql.Literal(supplier)
        )
    # print(supplier_filter)
    return supplier_filter

def div_filter_check(division: str):
    division = division.lower().split(",")
    if "full_fleet" in division:
        division_filter = sql.SQL("lower(division)")

    else:
        division_filter = sql.SQL("any({division})").format(
            division=sql.Literal(division)
        )

    return division_filter

def div_branch_filter_check(division: str, branch: str):
    division = division.lower().split(",")
    branch = branch.lower().split(",")
    if "full_fleet" in division:
        division_filter = sql.SQL("lower(division)")

    else:
        division_filter = sql.SQL("any({division})").format(
            division=sql.Literal(division)
        )

    if "all_branches" in branch:
        branches_filter = sql.SQL("lower(branch)")

    else:
        branches_filter = sql.SQL("any({branch})").format(branch=sql.Literal(branch))

    return division_filter, branches_filter

# // if we need to check div, branch and veh_type
def div_branch_type_filter_check(division: str, branch: str, veh_type: str):
    division = division.lower().split(",")
    if "full_fleet" in division:
        division_filter = sql.SQL("lower(division)")

    else:
        division_filter = sql.SQL("any({division})").format(
            division=sql.Literal(division)
        )

    branch = branch.lower().split(",")
    if "all_branches" in branch:
        branches_filter = sql.SQL("lower(branch)")

    else:
        branches_filter = sql.SQL("any({branch})").format(branch=sql.Literal(branch))

    veh_type = veh_type.lower().split(",")
    if "all_vehicles" in veh_type:
        type_filter = sql.SQL("lower(veh_type_map)")

    else:
        type_filter = sql.SQL("any({veh_type_map})").format(
            veh_type_map=sql.Literal(veh_type)
        )

    return division_filter, branches_filter, type_filter


## if we need to check component_map( component_dropdown)
def component_filter_check(component_map: str):
    component_map = component_map.lower().split(",")
    if "all_components" in component_map:
        component_filter = sql.SQL("lower(mapping)")

    else:
        component_filter = sql.SQL("any({component_map})").format(
            component_map=sql.Literal(component_map)
        )
        # print(component_filter)

    return component_filter

def date_filter(from_date: str, to_date: str):
    date_condition = sql.SQL("BETWEEN {from_date} AND {to_date}").format(
        from_date=sql.Literal(from_date), to_date=sql.Literal(to_date)
    )

    return date_condition

    # // if we need to check div, branch and veh_type


def div_branch_type_mapping_filter_check(
    division: str, branch: str, veh_type: str, mapping: str
):
    ## filtering by division, branch, veh_type, mapping
    division = division.lower().split(",")
    if "full_fleet" in division:
        division_filter = sql.SQL("lower(division)")

    else:
        division_filter = sql.SQL("any({division})").format(
            division=sql.Literal(division)
        )

    branch = branch.lower().split(",")
    if "all_branches" in branch:
        branches_filter = sql.SQL("lower(branch)")

    else:
        branches_filter = sql.SQL("any({branch})").format(branch=sql.Literal(branch))

    veh_type = veh_type.lower().split(",")
    if "all_vehicles" in veh_type:
        type_filter = sql.SQL("lower(veh_type_map)")

    else:
        type_filter = sql.SQL("any({veh_type_map})").format(
            veh_type_map=sql.Literal(veh_type)
        )

    mapping = mapping.lower().split(",")
    if "all_components" in mapping:
        component_filter = sql.SQL("lower(mapping)")

    else:
        component_filter = sql.SQL("any({mapping})").format(
            mapping=sql.Literal(mapping)
        )

    return division_filter, branches_filter, type_filter, component_filter




fnb_colours = [
    "#7c878e",
    "#3c474e",
    "#f2bc47",
    "#f39200",
    "#cf3f27",
    "#69d2dc",
    "#15a3b2",
    "#007582",
]


###used in monthly orders
def genrate_repair_count_df(df):
    ### get the repair counts and values
    registration_totals = (
        df.groupby(["vehiclereg", "fleet_no"])
        .agg(total_count=("vehiclereg", "size"), total_amount=("amount", "sum"))
        .reset_index()
    )

    # Step 2: Group by both 'vehiclereg' and 'mapping'
    mapping_details = (
        df.groupby(["vehiclereg", "mapping"])
        .agg(mapping_count=("mapping", "size"), mapping_amount=("amount", "sum"))
        .reset_index()
    )

    # Step 3: Pivot 'mapping_details' to have a column for each unique 'mapping' for count and amount
    mapping_pivot = mapping_details.pivot_table(
        index="vehiclereg",
        columns="mapping",
        values=["mapping_count", "mapping_amount"],
        fill_value=0,
    )

    # Rename and reorder the columns
    mapping_pivot.columns = [
        f'{col[1]} {col[0].split("_")[1]}' for col in mapping_pivot.columns
    ]
    mapping_pivot = mapping_pivot.reindex(
        sorted(
            mapping_pivot.columns,
            key=lambda x: (x.split()[0], x.split()[1] == "amount"),
        ),
        axis=1,
    )

    # Reset index to turn 'registration' back into a column
    mapping_pivot.reset_index(inplace=True)

    # Step 4: Join 'registration_totals' with 'mapping_pivot'
    repair_count_df = (
        pd.merge(registration_totals, mapping_pivot, on="vehiclereg", how="left")
        .sort_values(by="total_amount", ascending=False)
        .reset_index(drop=True)
    )
    return repair_count_df

#### WRITE THE FORM FILTER PARAMS TO THE EXCEL SHEET ####

def write_filter_parameters(workbook, sheet_name, startRow, formValues):
    ws = workbook[sheet_name]
    # ws.cell(row=startRow, column=1).value = formValues["julFromDate"]
    # ws.cell(row=startRow, column=2).value = "to"
    # ws.cell(row=startRow, column=3).value = formValues["julToDate"]
    startRow += 1
    ws.cell(row=startRow, column=1).value = "Division"
    ws.cell(row=startRow, column=2).value = "Branch"
    ws.cell(row=startRow, column=3).value = "Vehicle Type"
    ws.cell(row=startRow, column=4).value = "Vehicle Model"
    ws.cell(row=startRow, column=5).value = "Component"
    ws.cell(row=startRow, column=6).value = "Supplier"
    ws.cell(row=startRow, column=7).value = "Registration"
    startRow += 1
    writeRow = startRow
    for division in formValues["division"]:
        ws.cell(row=writeRow, column=1).value = division
        writeRow += 1
    writeRow = startRow
    for division in formValues["branch"]:
        ws.cell(row=writeRow, column=2).value = division
        writeRow += 1
    writeRow = startRow
    for division in formValues["vehicleType"]:
        ws.cell(row=writeRow, column=3).value = division
        writeRow += 1
    writeRow = startRow
    for division in formValues["models"]:
        ws.cell(row=writeRow, column=4).value = division
        writeRow += 1
    writeRow = startRow
    for division in formValues["components"]:
        ws.cell(row=writeRow, column=5).value = division
        writeRow += 1
    writeRow = startRow
    for division in formValues["suppliers"]:
        ws.cell(row=writeRow, column=6).value = division
        writeRow += 1
    writeRow = startRow
    for division in list(map(lambda x: x["vehiclereg"], formValues["registrations"])):
        ws.cell(row=writeRow, column=7).value = division
        writeRow += 1

######### SET COLUMN WIDTHS AUTOMATICALLY IN ALL SHEETS OF THE GIVEN WB #########
def set_column_widths(workbook):
    ##add the wb logo
    # logo_path = "/app/app/reports_templates/WesBank Logo Colour.png"

    # # Create an image object
    # img = Image(logo_path)

    # # Define the cell location where you want the top-left corner of the image to be
    # cell_location = "A1"

    # Add the image to the worksheet at the specified cell location
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]

        # Temporarily store images and then remove them
        images = ws._images
        ws._images = []

        # Resize columns (your existing code)
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2  # Adding a little extra space
            ws.column_dimensions[
                get_column_letter(column[0].column)
            ].width = adjusted_width

        # Re-add images
        for img in images:
            # You might need to adjust the properties of the images here
            ws.add_image(img)


def write_to_file_accrual_report(
    from_date: str, to_date: str, division: str, branch: str, veh_type: str
):
    con = return_connection()

    pivot_service_provider = f"""SELECT            
    serviceprovider,
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 1 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Jan 2023",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 2 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Feb 2023",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 3 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Mar 2023",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 4 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Apr 2023",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 5 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "May 2023",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 6 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Jun 2023",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 7 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Jul 2023",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 8 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Aug 2023",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 9 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Sep 2023",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 10 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Oct 2023",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 11 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Nov 2023",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 12 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Dec 2023",
    SUM(amount) AS "Total"
    FROM fleet.maintenance
        GROUP BY 
            ROLLUP (serviceprovider)
        ORDER BY 
    serviceprovider"""

    shp_query = f"""SELECT * FROM fleet.maintenance where 
    lower(division) = '{division}' and lower(branch) = '{branch}' and 
    lower(veh_type_map) = '{veh_type} 'and transdate between '{from_date}' and '{to_date}'
      """
    pivot_branch = f"""SELECT            
    branch,
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 1 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Jan 2023",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 2 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Feb 2023",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 3 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Mar 2023",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 4 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Apr 2023",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 5 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "May 2023",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 6 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Jun 2023",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 7 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Jul 2023",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 8 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Aug 2023",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 9 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Sep 2023",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 10 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Oct 2023",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 11 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Nov 2023",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 12 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Dec 2023",
    SUM(amount) AS "Total"
        FROM fleet.maintenance
        GROUP BY 
            ROLLUP (branch)
        ORDER BY 
    branch"""

    comparison_query = f"""	
      SELECT            
        veh_type_map,
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 1 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Jan 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 2 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Feb 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 3 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Mar 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 4 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Apr 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 5 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "May 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 6 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Jun 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 7 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Jul 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 8 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Aug 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 9 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Sep 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 10 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Oct 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 11 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Nov 2023",
        SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 12 AND EXTRACT(YEAR FROM transdate) = 2023 THEN amount ELSE 0 END) AS "Dec 2023",
        SUM(amount) AS "Total"
    FROM fleet.maintenance
    WHERE
    	veh_type_map in ('Passenger Car', 'Medium Commercial Vehicle','Light Commercial Vehicle', 'Heavy Commercial Vehicle' )

    GROUP BY 
        ROLLUP (veh_type_map) 
    ORDER BY 
        veh_type_map"""

    comparison_query_diff = f"""select	
          veh_type_map,
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 1 AND EXTRACT(YEAR FROM transdate) = 2022 THEN amount ELSE 0 END) AS "Jan 2022",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 2 AND EXTRACT(YEAR FROM transdate) = 2022 THEN amount ELSE 0 END) AS "Feb 2022",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 3 AND EXTRACT(YEAR FROM transdate) = 2022 THEN amount ELSE 0 END) AS "Mar 2022",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 4 AND EXTRACT(YEAR FROM transdate) = 2022 THEN amount ELSE 0 END) AS "Apr 2022",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 5 AND EXTRACT(YEAR FROM transdate) = 2022 THEN amount ELSE 0 END) AS "May 2022",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 6 AND EXTRACT(YEAR FROM transdate) = 2022 THEN amount ELSE 0 END) AS "Jun 2022",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 7 AND EXTRACT(YEAR FROM transdate) = 2022 THEN amount ELSE 0 END) AS "Jul 2022",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 8 AND EXTRACT(YEAR FROM transdate) = 2022 THEN amount ELSE 0 END) AS "Aug 2022",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 9 AND EXTRACT(YEAR FROM transdate) = 2022 THEN amount ELSE 0 END) AS "Sep 2022",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 10 AND EXTRACT(YEAR FROM transdate) = 2022 THEN amount ELSE 0 END) AS "Oct 2022",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 11 AND EXTRACT(YEAR FROM transdate) = 2022 THEN amount ELSE 0 END) AS "Nov 2022",
    SUM(CASE WHEN EXTRACT(MONTH FROM transdate) = 12 AND EXTRACT(YEAR FROM transdate) = 2022 THEN amount ELSE 0 END) AS "Dec 2022",
    SUM(amount) AS "Total"
    FROM fleet.maintenance
    WHERE
    	veh_type_map in ('Passenger Car', 'Medium Commercial Vehicle','Light Commercial Vehicle', 'Heavy Commercial Vehicle' )

    GROUP BY 
        ROLLUP (veh_type_map) 
    ORDER BY 
        veh_type_map"""

    comparison_query = pd.read_sql_query(
        comparison_query, con
    )  ## comparision query (passenger + commercial)
    comparison_query_diff = pd.read_sql_query(
        comparison_query_diff, con
    )  ## comparision query (passenger + commercial)

    pivot_sheet2 = pd.read_sql_query(pivot_branch, con)  ## branches
    pivot_sheet = pd.read_sql_query(pivot_service_provider, con)  ### servive providers
    ## shoprite transactions orders
    data_sheet4 = pd.read_sql_query(shp_query, con)
    ### accrual orders

    try:
        template_path = "/app/app/reports_templates/accrual_report_template.xlsx"
        workbook = openpyxl.load_workbook(template_path)

        sheet_name = "Data"
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")

        start_row = 2
        start_column = 1

        for col_num, header in enumerate(data_sheet4.columns, start_column):
            sheet.cell(row=2, column=col_num, value=header)

        for row_num, (_, row_data) in enumerate(data_sheet4.iterrows(), start_row):
            for col_num, header in enumerate(data_sheet4.columns, start_column):
                value = row_data[header]
                sheet.cell(row=row_num, column=col_num, value=value)

        workbook.save(template_path)
        # print("Data written to spreadsheet successfully!")

        template_path = "/app/app/reports_templates/accrual_report_template.xlsx"
        workbook = openpyxl.load_workbook(template_path)

        sheet_name = "Pivot"
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")

        start_row = 3
        start_column = 1

        for col_num, header in enumerate(pivot_sheet.columns, start_column):
            sheet.cell(row=2, column=col_num, value=header)

        for row_num, (_, row_data) in enumerate(pivot_sheet.iterrows(), start_row):
            for col_num, header in enumerate(pivot_sheet.columns, start_column):
                value = row_data[header]
                sheet.cell(row=row_num, column=col_num, value=value)

        workbook.save(template_path)
        # print("Data written to spreadsheet successfully!")

        template_path = "/app/app/reports_templates/accrual_report_template.xlsx"
        workbook = openpyxl.load_workbook(template_path)

        sheet_name = "Pivot"
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")

        start_row = 3
        start_column = 17

        start_column = 17

        for col_num, header in enumerate(pivot_sheet2.columns, start_column):
            sheet.cell(row=2, column=col_num, value=header)

        for row_num, (_, row_data) in enumerate(pivot_sheet2.iterrows(), start_row):
            for col_num, header in enumerate(pivot_sheet2.columns, start_column):
                value = row_data[header]
                sheet.cell(row=row_num, column=col_num, value=value)

        workbook.save(template_path)
        # print("Data written to spreadsheet successfully!")

        template_path = "/app/app/reports_templates/accrual_report_template.xlsx"
        workbook = openpyxl.load_workbook(template_path)

        sheet_name = "Comparison"
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")

        start_row = 2
        start_column = 1

        start_column = 1

        for col_num, header in enumerate(comparison_query.columns, start_column):
            sheet.cell(row=2, column=col_num, value=header)

        for row_num, (_, row_data) in enumerate(comparison_query.iterrows(), start_row):
            for col_num, header in enumerate(comparison_query.columns, start_column):
                value = row_data[header]
                sheet.cell(row=row_num, column=col_num, value=value)

        workbook.save(template_path)
        # print("Data written to spreadsheet successfully!")

        template_path = "/app/app/reports_templates/accrual_report_template.xlsx"
        workbook = openpyxl.load_workbook(template_path)

        sheet_name = "Comparison"
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")

        start_row = 14
        start_column = 1

        start_column = 1

        for col_num, header in enumerate(comparison_query_diff.columns, start_column):
            sheet.cell(row=2, column=col_num, value=header)

        for row_num, (_, row_data) in enumerate(
            comparison_query_diff.iterrows(), start_row
        ):
            for col_num, header in enumerate(
                comparison_query_diff.columns, start_column
            ):
                value = row_data[header]
                sheet.cell(row=row_num, column=col_num, value=value)

        workbook.save(template_path)
        # print("Data written to spreadsheet successfully!")

    except Exception as e:
        print(f"Error: {e}")
        raise

def julian_date_range(date):
    query = sql.SQL(
        """COPY(
                    SELECT jul_from_date, jul_to_date FROM fleet.julian_cal 
                    where {date} between jul_from_date and jul_to_date
    )TO STDOUT WITH CSV HEADER"""
    ).format(date=sql.Literal(date))
    res = exc_qrs_get_dfs_raw([query])[0].to_dict("records")
    from_date = res[0]["jul_from_date"]
    to_date = res[0]["jul_to_date"]

    return from_date, to_date

def get_financial_years(date_string):
    try:
        input_date = datetime.strptime(date_string, "%Y-%m-%d")
        current_year_start = datetime(input_date.year, 4, 1)

        # Calculate the start and end dates for the current financial year
        if input_date < current_year_start:
            current_year_start = datetime(input_date.year - 1, 4, 1)
            current_year_end = datetime(input_date.year, 3, 31)
        else:
            current_year_end = datetime(input_date.year + 1, 3, 31)

        # Calculate the start and end dates for the previous 2 financial years
        previous_year2_start = datetime(input_date.year - 2, 4, 1)
        previous_year2_end = datetime(input_date.year - 1, 3, 31)

        previous_year1_start = datetime(input_date.year - 1, 4, 1)
        previous_year1_end = datetime(input_date.year, 3, 31)
    except Exception as e:
        print(f"An error occurred: {e}")
        raise HTTPException(status_code=500, detail=f"Error Details: {traceback.format_exc()}")

    return {
        "current_financial_year": {
            "start_date": current_year_start.strftime("%Y-%m-%d"),
            "end_date": current_year_end.strftime("%Y-%m-%d"),
        },
        "previous_financial_year_1": {
            "start_date": previous_year1_start.strftime("%Y-%m-%d"),
            "end_date": previous_year1_end.strftime("%Y-%m-%d"),
        },
        "previous_financial_year_2": {
            "start_date": previous_year2_start.strftime("%Y-%m-%d"),
            "end_date": previous_year2_end.strftime("%Y-%m-%d"),
        },
    }


def fin_year_start(date):
    # Returns the start of the financial year for a given date in YYYY-MM-DD format
    if 7 <= date.month:
        year_start = date.replace(month=7)
    else:
        year_start = date.replace(year=date.year - 1, month=7)
    return year_start.strftime("%Y-%m-%d")

def fin_year_check(date):
    if 7 <= date.month <= 12:
        new_date = date.replace(year=date.year + 1)
        return new_date
    else:
        return date
def put_dates_in_dumb_format(df):
    if "quote_date" in df.columns:
        df["quote_date"] = pd.to_datetime(df["quote_date"]).dt.strftime("%d.%m.%Y")
    if "order_date" in df.columns:
        df["order_date"] = pd.to_datetime(df["order_date"]).dt.strftime("%d.%m.%Y")
    if "date" in df.columns:
        df["date"] = pd.to_datetime(df["date"]).dt.strftime("%d.%m.%Y")

    return df


def months_span(date_str1, date_str2):
    '''Calculates the number of months over two given dates'''
    date_format = "%Y-%m-%d"
    
    date1 = datetime.strptime(date_str1, date_format)
    date2 = datetime.strptime(date_str2, date_format)

    month_span = 1 + (date2.year - date1.year) * 12 + (date2.month - date1.month)

    return month_span

def last_twelve_months():
    '''Returns last 12 months from current day. Used for dynamic sql generation.'''
    end_date = datetime.now()
    start_date = end_date - timedelta(days=365)
    current = start_date

    months = []
    while current <= end_date:
        month_str = current.strftime("%Y-%m")
        if month_str not in months:
            months.append(month_str)
        current += timedelta(days=32)
        current = current.replace(day=1)

    return months

def calculate_boxplot_stats(dataframe, name: str, value: str):
    """
    Calculate box plot statistics for each service provider in the given DataFrame.

    Parameters:
    - dataframe (pd.DataFrame): The input DataFrame containing 'serviceprovider' and 'amount' columns.
    - name: The column containing names/labels.
    - value: The column containing values.

    Returns:
    - list of dict: A list of dictionaries with box plot statistics for each service provider.
    """

    grouped_data = dataframe.groupby(name)[value]
    
    boxplot_stats = grouped_data.describe(percentiles=[.25, .5, .75])

    result_list = []
    for name, stats in boxplot_stats.iterrows():
        result_dict = {
            'name': name,
            'min': stats['min'],
            'Q1': stats['25%'],
            'median': stats['50%'],
            'Q3': stats['75%'],
            'max': stats['max']
        }
        result_list.append(result_dict)

    return result_list
